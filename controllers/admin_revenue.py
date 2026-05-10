import re
import io
from collections import defaultdict
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
import openpyxl

from database import Base, engine, get_db
from models.admin_revenue import AdminRevenueRow

# Create table if it doesn't exist
Base.metadata.create_all(bind=engine, tables=[AdminRevenueRow.__table__])

router = APIRouter()

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_COLS = ["m1","m2","m3","m4","m5","m6","m7","m8","m9","m10","m11","m12"]

CLIENT_RE = re.compile(r"^(.+?)\s+\d{2,}\s+\d+\s+\d{2}/\d{2}/\d{2}")


def _rows_to_response(db_rows):
    rows = []
    for r in db_rows:
        months = [getattr(r, c) for c in MONTH_COLS]
        rows.append({"client": r.client, "months": months, "total": r.total})
    filename = db_rows[0].filename if db_rows else ""
    return {"months": MONTHS, "rows": rows, "filename": filename}


@router.get("/admin-revenue")
def get_admin_revenue(db: Session = Depends(get_db)):
    rows = db.query(AdminRevenueRow).order_by(AdminRevenueRow.client).all()
    if not rows:
        return {"months": MONTHS, "rows": [], "filename": ""}
    return _rows_to_response(rows)


@router.post("/admin-revenue/upload")
async def upload_admin_revenue(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)

    # Find the sheet with the GL data header
    ws = None
    for name in wb.sheetnames:
        s = wb[name]
        for row in s.iter_rows(min_row=1, max_row=10, values_only=True):
            if row and row[0] == "Posted dt.":
                ws = s
                break
        if ws:
            break
    if ws is None:
        raise HTTPException(status_code=422, detail="No sheet with 'Posted dt.' header found")

    # Locate header row
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        if row and row[0] == "Posted dt.":
            header_row = i
            break

    # Aggregate credits by client and month
    totals = defaultdict(lambda: [0.0] * 12)
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        posted_dt, _, _, memo, _, _, _, credit, *_ = (list(row) + [None]*9)[:9]
        if not posted_dt or not memo or not credit:
            continue
        if not hasattr(posted_dt, "month"):
            continue
        m = CLIENT_RE.match(str(memo).strip())
        if not m:
            continue
        totals[m.group(1).strip()][posted_dt.month - 1] += float(credit)

    # Replace all rows in DB
    db.query(AdminRevenueRow).delete()
    for client, monthly in sorted(totals.items()):
        kwargs = {c: round(v, 2) for c, v in zip(MONTH_COLS, monthly)}
        db.add(AdminRevenueRow(
            client=client,
            filename=file.filename,
            total=round(sum(monthly), 2),
            **kwargs,
        ))
    db.commit()

    db_rows = db.query(AdminRevenueRow).order_by(AdminRevenueRow.client).all()
    return _rows_to_response(db_rows)
