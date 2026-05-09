"""
Import WC Rates, WC Guidelines, and SUTA Cost Rates from the Excel pricing template.
Idempotent — deletes and re-inserts all rows each run.

Run from the project root:
    python testing/import_rates.py
"""

import sys
import os
import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from database import SessionLocal
from models import WCRate, WCGuideline, SutaRate

EXCEL_PATH = r"C:\workspaces\business\vested-hr\documents\Pricing Template 03.10.26_ALL SHEETS.xlsx"


def _str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _date_str(val):
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).strip() or None


def import_wc_rates(ws, db):
    db.query(WCRate).delete()
    count = 0
    row_id = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        carrier = _str(row[0])
        state = _str(row[1])
        class_code = _str(row[2])
        concat = _str(row[3])
        rate = _float(row[4])
        effective_date = _date_str(row[5])
        min_premium = _float(row[6])
        description = _str(row[7])

        if not state or not class_code:
            continue
        if concat is None:
            concat = (state + class_code) if state and class_code else None

        db.add(WCRate(
            id=row_id,
            carrier=carrier,
            state=state,
            class_code=class_code,
            concat=concat,
            rate=rate,
            min_premium=min_premium,
            description=description,
            effective_date=effective_date,
        ))
        row_id += 1
        count += 1

    db.commit()
    return count


def import_wc_guidelines(ws, db):
    db.query(WCGuideline).delete()
    count = 0
    row_id = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        state = _str(row[0])
        ncci_code = _str(row[1])
        lookup_code = _str(row[2])
        concat = _str(row[3])
        irmi_classification = _str(row[4])
        naics = _str(row[5])
        hazard_group = _str(row[6])
        flag_100k = _str(row[7])
        effective_date = _date_str(row[8])

        if not state or not ncci_code:
            continue
        if concat is None:
            concat = (state + ncci_code) if state and ncci_code else None

        db.add(WCGuideline(
            id=row_id,
            state=state,
            ncci_code=ncci_code,
            lookup_code=lookup_code,
            concat=concat,
            irmi_classification=irmi_classification,
            naics=naics,
            hazard_group=hazard_group,
            flag_100k=flag_100k,
            effective_date=effective_date,
        ))
        row_id += 1
        count += 1

    db.commit()
    return count


def import_suta_rates(ws, db):
    db.query(SutaRate).delete()
    seen_states = set()
    count = 0
    row_id = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        state = _str(row[0])
        if not state or len(state) != 2 or state in seen_states:
            continue

        threshold = _float(row[1])
        raw_vhr = _float(row[2])
        # Sheet stores rates as percentages (2.7 = 2.7%); DB stores decimals (0.027)
        vhr_min_rate = raw_vhr / 100 if raw_vhr is not None else None

        cr_raw = _str(row[3])
        client_reporting = cr_raw == "Y" if cr_raw else False

        date_updated = _date_str(row[4])

        # Our Cost is in column J (index 9)
        raw_cost = _float(row[9])
        our_cost = raw_cost / 100 if raw_cost is not None else None

        db.add(SutaRate(
            id=row_id,
            state=state,
            threshold=threshold,
            vhr_min_rate=vhr_min_rate,
            client_reporting=client_reporting,
            our_cost=our_cost,
            date_updated=date_updated,
        ))
        seen_states.add(state)
        row_id += 1
        count += 1

    db.commit()
    return count


def main():
    print(f"Loading workbook from {EXCEL_PATH} ...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    db = SessionLocal()
    try:
        print("Importing WC Cost Rates ...")
        n = import_wc_rates(wb["WC Cost Rates"], db)
        print(f"  {n} rows imported")

        print("Importing WC Sunz Guidelines ...")
        n = import_wc_guidelines(wb["WC Sunz Guidelines"], db)
        print(f"  {n} rows imported")

        print("Importing SUTA Cost Rates ...")
        n = import_suta_rates(wb["SUTA Cost Rates"], db)
        print(f"  {n} rows imported")
    finally:
        db.close()

    print("Done.")


if __name__ == "__main__":
    main()
